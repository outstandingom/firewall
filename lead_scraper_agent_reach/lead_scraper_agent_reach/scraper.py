"""
╔══════════════════════════════════════════════════════════╗
║     LEAD SCRAPER WITH AGENT-REACH INTEGRATION            ║
║  Uses Exa for search + Jina Reader for web content       ║
╚══════════════════════════════════════════════════════════╝

Just run: python scraper.py

This script uses Agent-Reach's:
  - Exa for semantic web search (free)
  - Jina Reader for clean web page extraction
"""

import re
import time
import sys
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# ══════════════════════════════════════════════
#  ★  CHANGE YOUR SEARCH QUERY HERE  ★
# ══════════════════════════════════════════════
SEARCH_QUERY  = "education institutes in Bhopal"  # ← EDIT THIS
NUM_RESULTS   = 25
DELAY_SECONDS = 1
OUTPUT_FILE   = "leads_output.xlsx"
# ══════════════════════════════════════════════

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
}

EMAIL_PATTERN = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
)
PHONE_PATTERN = re.compile(
    r"(?:"
    r"\+91[\s\-\.]?\d{5}[\s\-\.]?\d{5}"
    r"|\+91[\s\-]?\d{10}"
    r"|\b0\d{2,4}[\s\-]\d{6,8}\b"
    r"|\b0\d{9,10}\b"
    r"|\b[6-9]\d{9}\b"
    r")"
)

SKIP_EMAIL_DOMAINS = {
    "example.com", "yourdomain.com", "domain.com", "email.com",
    "sentry.io", "wixpress.com", "schema.org", "w3.org",
}


def clean_emails(emails):
    result = []
    for e in emails:
        domain = e.split("@")[-1].lower()
        if domain not in SKIP_EMAIL_DOMAINS and len(e) <= 80:
            result.append(e.lower())
    return list(dict.fromkeys(result))


def search_with_exa(query: str, num: int) -> list:
    print("   🚀 Using Yahoo Search (Free Fallback)...")
    try:
        import urllib.request
        import urllib.parse
        import re
        
        q = urllib.parse.quote(query)
        req = urllib.request.Request(
            f'https://search.yahoo.com/search?p={q}&n={num}',
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        )
        html = urllib.request.urlopen(req).read().decode('utf-8', errors='ignore')
        
        raw_urls = re.findall(r'RU=([^/]+)/RK=2', html)
        urls = []
        for u in raw_urls:
            decoded = urllib.parse.unquote(u)
            if 'yahoo.com' not in decoded and decoded not in urls:
                urls.append(decoded)
                if len(urls) >= num:
                    break
                    
        print(f"   ✅ Found {len(urls)} results!")
        return urls
    except Exception as e:
        print(f"   ⚠️  Yahoo Search failed: {e}")
        return []


def read_page_with_jina(url: str) -> str:
    try:
        jina_url = f"https://r.jina.ai/{url}"
        resp = requests.get(jina_url, headers=HEADERS, timeout=10)
        if resp.status_code == 200:
            return resp.text
    except Exception:
        pass
    
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10, allow_redirects=True)
        if resp.status_code == 200:
            return resp.text
    except Exception:
        pass
    
    return ""


def extract_from_content(url: str, content: str) -> dict:
    if not content:
        return {}

    if "<html" in content.lower():
        soup = BeautifulSoup(content, "html.parser")
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()
        text = soup.get_text(separator=" ", strip=True)
        title = soup.title.string if soup.title else ""
    else:
        text = content
        title = ""

    emails = clean_emails(EMAIL_PATTERN.findall(text))
    phones = list(dict.fromkeys(PHONE_PATTERN.findall(text)))
    phones = [p.strip() for p in phones]

    if not title:
        lines = text.split("\n")
        for line in lines[:5]:
            if 10 < len(line) < 150:
                title = line
                break

    description = ""
    sentences = re.split(r'[.!?\n]', text)
    for sent in sentences:
        sent = sent.strip()
        if 50 < len(sent) < 300:
            description = sent
            break

    return {
        "emails": emails[:8],
        "phones": phones[:8],
        "title": title[:100] if title else "",
        "description": description[:300] if description else "",
    }


def scrape_with_agent_reach(url: str) -> dict:
    result = {
        "Website Name": "",
        "URL": url,
        "Emails": "",
        "Phone Numbers": "",
        "Description": "",
        "Method": "",
        "Email Count": 0,
        "Phone Count": 0,
        "Status": "",
    }

    content = read_page_with_jina(url)
    if not content:
        result["Status"] = "Could not load"
        result["Method"] = "Failed"
        return result

    info = extract_from_content(url, content)
    all_emails = list(info.get("emails", []))
    all_phones = list(info.get("phones", []))

    result["Website Name"] = info.get("title", "")
    result["Description"] = info.get("description", "")
    result["Method"] = "Jina Reader"

    if not all_emails and not all_phones:
        for suffix in ["/contact", "/contact-us", "/contactus"]:
            contact_url = url.rstrip("/") + suffix
            contact_content = read_page_with_jina(contact_url)
            if contact_content:
                info2 = extract_from_content(contact_url, contact_content)
                if info2.get("emails") or info2.get("phones"):
                    all_emails.extend(info2.get("emails", []))
                    all_phones.extend(info2.get("phones", []))
                    result["Method"] = f"Jina (found on {suffix})"
                    break
            time.sleep(0.5)

    all_emails = list(dict.fromkeys(all_emails))[:8]
    all_phones = list(dict.fromkeys(all_phones))[:8]

    result["Emails"] = ", ".join(all_emails)
    result["Phone Numbers"] = ", ".join(all_phones)
    result["Email Count"] = len(all_emails)
    result["Phone Count"] = len(all_phones)
    result["Status"] = "OK"
    return result


def save_to_excel(data: list, filename: str, query: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    COLS = [
        "Website Name", "URL", "Emails", "Phone Numbers",
        "Description", "Method", "Email Count", "Phone Count", "Status"
    ]
    COL_WIDTHS = [35, 45, 55, 35, 55, 20, 12, 12, 12]

    hdr_fill = PatternFill("solid", fgColor="1A3E6C")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hit_fill = PatternFill("solid", fgColor="E3F2FD")
    good_fill = PatternFill("solid", fgColor="C8E6C9")

    for ci, (col, w) in enumerate(zip(COLS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 28

    for ri, row in enumerate(data, 2):
        has_email = row.get("Email Count", 0) > 0
        has_phone = row.get("Phone Count", 0) > 0
        fill = good_fill if (has_email and has_phone) else (hit_fill if (has_email or has_phone) else None)
        for ci, col in enumerate(COLS, 1):
            val = row.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill
            if col == "URL" and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
        ws.row_dimensions[ri].height = 40

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Summary")
    total = len(data)
    ok = sum(1 for d in data if d.get("Status") == "OK")
    with_email = sum(1 for d in data if d.get("Email Count", 0) > 0)
    with_phone = sum(1 for d in data if d.get("Phone Count", 0) > 0)
    with_both = sum(1 for d in data if d.get("Email Count", 0) > 0 and d.get("Phone Count", 0) > 0)

    summary = [
        ("Search Query", query),
        ("Date", datetime.now().strftime("%d %b %Y  %H:%M")),
        ("Websites Checked", total),
        ("Pages Loaded OK", ok),
        ("━━━━━━━━", "━━━━━━━━━━━━━━━━━━━"),
        ("With Email", with_email),
        ("With Phone", with_phone),
        ("With BOTH", with_both),
        ("No Contact Info", total - with_email - with_phone + with_both),
    ]
    for ri, (k, v) in enumerate(summary, 1):
        ws2.cell(row=ri, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=ri, column=2, value=str(v))
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 45

    wb.save(filename)


def main():
    import warnings
    warnings.filterwarnings("ignore")

    query = SEARCH_QUERY
    if len(sys.argv) > 1:
        query = " ".join(sys.argv[1:])

    print("=" * 70)
    print("  LEAD SCRAPER WITH AGENT-REACH")
    print(f"  Query  : {query}")
    print(f"  Results: up to {NUM_RESULTS} websites")
    print("=" * 70)

    print("\n🔍 Searching...")
    urls = search_with_exa(query, NUM_RESULTS)

    if not urls:
        print("\n   ❌ Google search didn't work. Paste URLs manually:\n")
        urls = []
        while True:
            u = input("   Enter URL (or 'done' to finish): ").strip()
            if u.lower() in ("done", "exit", ""):
                break
            if u.startswith("http"):
                urls.append(u)
        if not urls:
            print("No URLs. Exiting.")
            return

    print(f"\n✅ Found {len(urls)} websites\n")

    results = []
    for i, url in enumerate(urls, 1):
        short = url[:65] + "..." if len(url) > 65 else url
        print(f"[{i:02d}/{len(urls)}] {short}")

        info = scrape_with_agent_reach(url)

        ec = info.get("Email Count", 0)
        pc = info.get("Phone Count", 0)
        icons = []
        if ec:
            icons.append(f"📧 {ec} email(s)")
        if pc:
            icons.append(f"📞 {pc} phone(s)")
        status_str = "  ".join(icons) if icons else "  (no contact)"
        print(f"         {status_str}   [{info.get('Method', 'N/A')}]")
        if info.get("Emails"):
            print(f"         → {info['Emails'][:70]}")
        if info.get("Phone Numbers"):
            print(f"         → {info['Phone Numbers'][:60]}")

        results.append(info)
        time.sleep(DELAY_SECONDS)

    save_to_excel(results, OUTPUT_FILE, query)

    with_any = [r for r in results if r.get("Email Count", 0) > 0 or r.get("Phone Count", 0) > 0]
    print(f"\n{'='*70}")
    print(f"  ✅ DONE!")
    print(f"  Websites checked   : {len(results)}")
    print(f"  With contact info  : {len(with_any)}")
    print(f"  File saved         : {OUTPUT_FILE}")
    print(f"{'='*70}")
    print(f"\n  Open '{OUTPUT_FILE}' in Excel to see results.")


if __name__ == "__main__":
    main()
