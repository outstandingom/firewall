"""
╔══════════════════════════════════════════════════════╗
║          LEAD GENERATION SCRAPER                     ║
║  Searches websites → Extracts Email, Phone, Info     ║
╚══════════════════════════════════════════════════════╝

HOW TO USE:
  1. Install Python from https://python.org
  2. Open terminal / command prompt
  3. Run: pip install requests beautifulsoup4 openpyxl googlesearch-python
  4. Edit SEARCH_QUERY below to your topic
  5. Run: python lead_scraper.py
  6. Open the generated leads_output.xlsx file

EXAMPLES of SEARCH_QUERY:
  "education institutes in Bhopal"
  "coaching classes in Bhopal"
  "private schools in Bhopal"
  "IT companies in Bhopal"
  "hospitals in Indore Madhya Pradesh"
"""

import re
import time
import sys
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# ══════════════════════════════════════════════
#  ★  CHANGE THESE SETTINGS  ★
# ══════════════════════════════════════════════
SEARCH_QUERY  = "education institutes in Bhopal"  # ← your search keyword
NUM_RESULTS   = 30                                  # how many websites to check
DELAY_SECONDS = 2                                   # pause between requests
OUTPUT_FILE   = "leads_output.xlsx"                 # output file name
# ══════════════════════════════════════════════

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

EMAIL_PATTERN = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
)
PHONE_PATTERN = re.compile(
    r"(?:"
    r"\+91[\s\-\.]?\d{5}[\s\-\.]?\d{5}"   # +91-XXXXX-XXXXX
    r"|\+91[\s\-]?\d{10}"                   # +91 XXXXXXXXXX
    r"|\b0\d{2,4}[\s\-]\d{6,8}\b"          # 0755-123456
    r"|\b0\d{9,10}\b"                       # 07554123456
    r"|\b[6-9]\d{9}\b"                      # 10-digit mobile
    r")"
)

SKIP_EMAIL_DOMAINS = {
    "example.com", "yourdomain.com", "domain.com", "email.com",
    "sentry.io", "wixpress.com", "schema.org", "w3.org",
    "placeholder.com", "test.com", "sample.com"
}


def clean_emails(emails):
    result = []
    for e in emails:
        domain = e.split("@")[-1].lower()
        if domain not in SKIP_EMAIL_DOMAINS and len(e) <= 80:
            result.append(e.lower())
    return list(dict.fromkeys(result))  # deduplicate preserving order


def search_google(query: str, num: int) -> list:
    """Search Google and return URLs."""
    urls = []
    try:
        from googlesearch import search as gsearch
        print("   Using googlesearch library...")
        for url in gsearch(query, num_results=num, lang="en", sleep_interval=1):
            if url and url.startswith("http") and "google" not in url:
                urls.append(url)
        return urls
    except ImportError:
        print("   googlesearch not installed, trying direct method...")
    except Exception as e:
        print(f"   googlesearch error ({e}), trying direct method...")

    # Fallback: direct Google scraping
    try:
        import urllib.parse
        encoded = urllib.parse.quote(query)
        resp = requests.get(
            f"https://www.google.com/search?q={encoded}&num={num+5}&hl=en",
            headers=HEADERS, timeout=12
        )
        soup = BeautifulSoup(resp.text, "html.parser")
        for a in soup.select("a"):
            href = a.get("href", "")
            if href.startswith("/url?q="):
                url = href.split("/url?q=")[1].split("&")[0]
                import urllib.parse as up
                url = up.unquote(url)
                if url.startswith("http") and "google" not in url:
                    urls.append(url)
        print(f"   Direct scraping found {len(urls)} URLs")
    except Exception as e:
        print(f"   Direct Google scraping failed: {e}")

    return list(dict.fromkeys(urls))[:num]


def fetch_page(url: str) -> str:
    """Download a page, trying contact/about pages too."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=12,
                            allow_redirects=True, verify=False)
        if resp.status_code == 200:
            return resp.text
    except Exception:
        pass
    return ""


def extract_from_page(url: str, html: str) -> dict:
    """Extract contact info from HTML."""
    if not html:
        return {}

    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript", "head"]):
        tag.decompose()

    text = soup.get_text(separator=" ", strip=True)

    emails = clean_emails(EMAIL_PATTERN.findall(text))
    phones = list(dict.fromkeys(PHONE_PATTERN.findall(text)))
    phones = [p.strip() for p in phones]

    title = ""
    if soup.title and soup.title.string:
        title = soup.title.string.strip()[:100]

    description = ""
    for meta_name in [{"name": "description"}, {"property": "og:description"}]:
        tag = soup.find("meta", attrs=meta_name)
        if tag and tag.get("content"):
            description = tag["content"].strip()[:300]
            break

    return {
        "emails": emails[:8],
        "phones": phones[:8],
        "title": title,
        "description": description,
    }


def scrape_website(url: str) -> dict:
    """Scrape a website: try main page, then contact page."""
    result = {
        "Website Name": "",
        "URL": url,
        "Emails": "",
        "Phone Numbers": "",
        "Description": "",
        "Address Hint": "",
        "Contact Page": "",
        "Email Count": 0,
        "Phone Count": 0,
        "Status": "",
    }

    # 1. Main page
    html = fetch_page(url)
    if not html:
        result["Status"] = "Could not load"
        return result

    info = extract_from_page(url, html)
    result["Website Name"]  = info.get("title", "")
    result["Description"]   = info.get("description", "")
    all_emails = list(info.get("emails", []))
    all_phones = list(info.get("phones", []))

    # 2. Also check contact page if main page has no contacts
    if not all_emails and not all_phones:
        base = url.rstrip("/")
        for suffix in ["/contact", "/contact-us", "/contactus", "/about", "/about-us"]:
            contact_url = base + suffix
            html2 = fetch_page(contact_url)
            if html2:
                info2 = extract_from_page(contact_url, html2)
                if info2.get("emails") or info2.get("phones"):
                    all_emails.extend(info2.get("emails", []))
                    all_phones.extend(info2.get("phones", []))
                    result["Contact Page"] = contact_url
                    break
            time.sleep(0.5)

    # Deduplicate
    all_emails = list(dict.fromkeys(all_emails))[:8]
    all_phones = list(dict.fromkeys(all_phones))[:8]

    result["Emails"]       = ", ".join(all_emails)
    result["Phone Numbers"] = ", ".join(all_phones)
    result["Email Count"]  = len(all_emails)
    result["Phone Count"]  = len(all_phones)
    result["Status"]       = "OK"
    return result


def save_to_excel(data: list, filename: str, query: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    COLS = [
        "Website Name", "URL", "Emails", "Phone Numbers",
        "Description", "Contact Page", "Email Count", "Phone Count", "Status"
    ]
    COL_WIDTHS = [35, 45, 55, 35, 55, 40, 12, 12, 12]

    hdr_fill = PatternFill("solid", fgColor="1A3E6C")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hit_fill = PatternFill("solid", fgColor="E3F2FD")
    good_fill= PatternFill("solid", fgColor="C8E6C9")

    for ci, (col, w) in enumerate(zip(COLS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 28

    for ri, row in enumerate(data, 2):
        has_email = row.get("Email Count", 0) > 0
        has_phone = row.get("Phone Count", 0) > 0
        fill = good_fill if (has_email and has_phone) else (hit_fill if (has_email or has_phone) else None)
        for ci, col in enumerate(COLS, 1):
            val = row.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill
            if col == "URL" and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
        ws.row_dimensions[ri].height = 40

    # Freeze top row + auto filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total      = len(data)
    ok         = sum(1 for d in data if d.get("Status") == "OK")
    with_email = sum(1 for d in data if d.get("Email Count", 0) > 0)
    with_phone = sum(1 for d in data if d.get("Phone Count", 0) > 0)
    with_both  = sum(1 for d in data if d.get("Email Count", 0) > 0 and d.get("Phone Count", 0) > 0)

    summary = [
        ("Search Query",     query),
        ("Date",             datetime.now().strftime("%d %b %Y  %H:%M")),
        ("Websites Checked", total),
        ("Pages Loaded OK",  ok),
        ("━━━━━━━━",         "━━━━━━━━━━━━━━━━━━━━━━━━"),
        ("With Email",       with_email),
        ("With Phone",       with_phone),
        ("With BOTH",        with_both),
        ("No Contact Info",  total - with_email - with_phone + with_both),
    ]
    for ri, (k, v) in enumerate(summary, 1):
        ws2.cell(row=ri, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=ri, column=2, value=str(v))
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 45

    wb.save(filename)


def main():
    import warnings
    warnings.filterwarnings("ignore")  # suppress SSL warnings

    query = SEARCH_QUERY
    if len(sys.argv) > 1:
        query = " ".join(sys.argv[1:])

    print("=" * 60)
    print("  LEAD GENERATION SCRAPER")
    print(f"  Query  : {query}")
    print(f"  Results: up to {NUM_RESULTS} websites")
    print("=" * 60)

    print("\n🔍 Searching Google...")
    urls = search_google(query, NUM_RESULTS)

    if not urls:
        print("\n❌ Could not get search results automatically.")
        print("   You can manually paste website URLs below.")
        print("   Type each URL and press Enter. Type 'done' when finished.\n")
        while True:
            u = input("   URL: ").strip()
            if u.lower() in ("done", "exit", ""):
                break
            if u.startswith("http"):
                urls.append(u)
        if not urls:
            print("No URLs provided. Exiting.")
            return

    print(f"\n✅ Found {len(urls)} websites to check\n")

    results = []
    for i, url in enumerate(urls, 1):
        short = url[:65] + "..." if len(url) > 65 else url
        print(f"[{i:02d}/{len(urls)}] {short}")

        info = scrape_website(url)

        ec = info.get("Email Count", 0)
        pc = info.get("Phone Count", 0)
        icons = []
        if ec: icons.append(f"📧 {ec} email(s)")
        if pc: icons.append(f"📞 {pc} phone(s)")
        status_str = "  ".join(icons) if icons else "  (no contact found)"
        print(f"         {status_str}")
        if info.get("Emails"):   print(f"         → {info['Emails'][:70]}")
        if info.get("Phone Numbers"): print(f"         → {info['Phone Numbers'][:60]}")

        results.append(info)
        time.sleep(DELAY_SECONDS)

    # Save
    save_to_excel(results, OUTPUT_FILE, query)

    # Summary
    with_any = [r for r in results if r.get("Email Count", 0) > 0 or r.get("Phone Count", 0) > 0]
    print(f"\n{'='*60}")
    print(f"  ✅ DONE!")
    print(f"  Websites checked   : {len(results)}")
    print(f"  With contact info  : {len(with_any)}")
    print(f"  File saved         : {OUTPUT_FILE}")
    print(f"{'='*60}")
    print(f"\n  Open '{OUTPUT_FILE}' in Excel to see all results.")


if __name__ == "__main__":
    main()